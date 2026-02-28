# -*- coding: utf-8 -*-
"""
excel_to_txt.py（openpyxl稳定版｜完整可一键复制粘贴）

修正点（针对你这类电视参数对比表）：
- 自动识别“型号”所在行作为“列头行”（每列一个机型：X11L / Q10M Ultra ...）
- 自动识别“产品定位”所在行（通常在“型号”上一行），并写入每个机型 block：产品定位=xxx
- 后续参数行按 “参数名=值” 输出
- 单元格里有多行/多个参数：合并为 "; " 输出
"""

from __future__ import annotations

import re
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


_SPLIT_ITEMS_RE = re.compile(r"[;\n\|；｜]+")


def _norm(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s.lower() == "nan":
        return ""
    return s


def _compact_cell(v) -> str:
    v = _norm(v)
    if not v:
        return ""
    parts = [p.strip() for p in _SPLIT_ITEMS_RE.split(v) if p.strip()]
    if not parts:
        return ""
    seen = set()
    out = []
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return "; ".join(out)


def detect_brand_from_filename(filename: str) -> str:
    name = (filename or "").lower()
    if "tcl" in name:
        return "TCL"
    if "vidda" in name:
        return "Vidda"
    if "海信" in filename:
        return "海信"
    if "创维" in filename:
        return "创维"
    if "雷鸟" in filename:
        return "雷鸟"
    if "小米" in filename:
        return "小米"
    return "未知品牌"


def _find_row_index(rows: List[List], key: str, search_limit: int = 40) -> Optional[int]:
    """
    在前 search_limit 行里找：第1列 == key（忽略空白）
    """
    key = key.strip()
    for i in range(min(search_limit, len(rows))):
        if not rows[i]:
            continue
        first = _norm(rows[i][0])
        if first == key:
            return i
    return None


def _sheet_to_products(ws, brand: str) -> Dict[str, List[str]]:
    """
    解析一个 sheet：
    - 找到 “型号” 行：这一行的第2列开始是机型列表（18个）
    - 找到 “产品定位” 行：通常是“型号”上一行（也可能在别处），第2列开始是定位列表
    - 从“型号”行的下一行开始，按参数输出
    """
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # 去掉全空行（但保留行序以便定位）
    # 这里不强行删除，避免行号错位；只在取值时跳过空参数即可

    if not rows:
        return {}

    model_row_idx = _find_row_index(rows, "型号")
    if model_row_idx is None:
        # 有些表可能用“机型/型号名称”等
        for alt in ["机型", "型号名称", "型号（机型）"]:
            model_row_idx = _find_row_index(rows, alt)
            if model_row_idx is not None:
                break

    if model_row_idx is None:
        return {}

    model_row = rows[model_row_idx]
    # 机型列：从第二列开始
    model_names = [_norm(x) for x in model_row[1:]]
    # 过滤空列
    valid_cols = []
    for ci, m in enumerate(model_names, start=1):
        if m:
            valid_cols.append((ci, m))  # (列索引, 型号)

    if len(valid_cols) < 2:
        # 少于2个机型通常说明没识别对
        return {}

    # 产品定位行：优先找 “产品定位”
    pos_row_idx = _find_row_index(rows, "产品定位")
    if pos_row_idx is None and model_row_idx > 0:
        # 有的表“产品定位”就在“型号”上一行，但第一列可能没写“产品定位”（比如合并/空）
        # 这里不做强推断，避免误判
        pass

    pos_map: Dict[int, str] = {}
    if pos_row_idx is not None:
        pos_row = rows[pos_row_idx]
        for ci, _m in valid_cols:
            pos_map[ci] = _compact_cell(pos_row[ci] if ci < len(pos_row) else "")

    # 从型号行下一行开始，读取参数
    products: Dict[str, List[str]] = {}
    for ci, model in valid_cols:
        lines: List[str] = []
        lines.append("[PRODUCT]")
        lines.append(f"品牌={brand}")
        lines.append(f"型号={model}")
        lines.append(f"来源Sheet={ws.title}")
        if pos_map.get(ci):
            lines.append(f"产品定位={pos_map[ci]}")
        lines.append("")

        # 参数行从 model_row_idx+1 往下
        for r in rows[model_row_idx + 1 :]:
            if not r or len(r) == 0:
                continue
            param = _norm(r[0])
            if not param:
                continue
            # 跳过重复的“型号/产品定位”行
            if param in ("型号", "产品定位"):
                continue

            val = r[ci] if ci < len(r) else None
            vv = _compact_cell(val)
            if not vv:
                continue
            lines.append(f"{param}={vv}")

        lines.append("")
        products[model] = lines

    return products


def excel_to_txt(
    excel_path: str,
    filename_for_brand: str,
    sheet: Optional[str] = None,
) -> Tuple[str, str]:
    brand = detect_brand_from_filename(filename_for_brand)
    wb = load_workbook(excel_path, data_only=True)

    sheet_names = [sheet] if sheet and sheet in wb.sheetnames else list(wb.sheetnames)

    blocks: List[str] = []
    blocks.append("# GENERATED_EXCEL_TO_TXT")
    blocks.append(f"# brand={brand}")
    blocks.append("")

    hit_any = False
    for sh in sheet_names:
        ws = wb[sh]
        products = _sheet_to_products(ws, brand=brand)
        if not products:
            continue

        hit_any = True
        blocks.append(f"# ===== sheet: {ws.title} =====")
        blocks.append("")
        for _model, lines in products.items():
            blocks.extend(lines)

    if not hit_any:
        blocks.append("# WARN: no compatible sheet detected.")
        blocks.append("# Need a row whose col1 is '型号' and model names start from col2.")
        blocks.append("")

    return brand, "\n".join(blocks)