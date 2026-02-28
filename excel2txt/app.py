from flask import Flask, render_template, request, send_file, jsonify
import os
import uuid
import csv
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cell_to_str(v):
    """
    单元格值转字符串，完整保留所有内容：
    - None / 纯空白 → 空字符串（不用--代替）
    - '--' 原样保留（这是表格作者填写的占位符，属于有效内容）
    - 数字去掉多余小数（如 2025.0 → 2025，2025.5 保留）
    - 单元格内换行符 \n 保留（输出时做特殊处理）
    """
    if v is None:
        return ''
    if isinstance(v, float):
        # 整数型浮点数去掉 .0
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v)

def read_excel(path, sheets_mode='all'):
    """
    读取 Excel，返回 {sheet_name: [[cell_str, ...], ...]}
    只保留有实际内容的行（至少一个非空单元格），去除末尾全空行。
    """
    wb = load_workbook(path, data_only=True)
    names = wb.sheetnames
    if sheets_mode == 'first':
        names = names[:1]
    result = {}
    for name in names:
        ws = wb[name]
        rows = []
        # 找出实际有数据的最大列数
        max_col = ws.max_column or 1
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                 min_col=1, max_col=max_col, values_only=True):
            cells = [cell_to_str(v) for v in row]
            # 跳过完全空白的行
            if any(c.strip() for c in cells):
                rows.append(cells)
        result[name] = rows
    wb.close()
    return result

def read_csv_file(path):
    rows = []
    encodings = ['utf-8-sig', 'gbk', 'utf-8', 'latin-1']
    for enc in encodings:
        try:
            with open(path, 'r', encoding=enc, newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    if any(v.strip() for v in row):
                        rows.append(row)
            break
        except (UnicodeDecodeError, Exception):
            continue
    return {'Sheet1': rows}

def rows_to_txt_lines(rows, sep, newline_replacement):
    """
    将二维单元格数组转成 txt 行列表。
    单元格内的换行符用 newline_replacement 替换，保证每行是一行。
    """
    lines = []
    for row in rows:
        # 单元格内的 \n 替换成指定符号，避免破坏行结构
        cells = [c.replace('\n', newline_replacement).replace('\r', '') for c in row]
        lines.append(sep.join(cells))
    return lines

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return jsonify({'error': '没有选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': '不支持的文件格式，请上传 .xlsx .xls 或 .csv 文件'}), 400

    try:
        filename = secure_filename(file.filename)
        uid = str(uuid.uuid4())[:8]
        upload_path = os.path.join(app.config['UPLOAD_FOLDER'], uid + '_' + filename)
        file.save(upload_path)

        ext = filename.rsplit('.', 1)[1].lower()
        separator    = request.form.get('separator', 'tab')
        sheets_mode  = request.form.get('sheets_mode', 'all')
        # 单元格内换行的替换方式
        newline_mode = request.form.get('newline_mode', 'slash')

        sep_map = {'tab': '\t', 'comma': ',', 'pipe': '|', 'space': ' '}
        sep = sep_map.get(separator, '\t')

        nl_map = {'slash': ' / ', 'space': ' ', 'keep': '\n', 'br': '↵'}
        nl_rep = nl_map.get(newline_mode, ' / ')

        if ext == 'csv':
            sheets = read_csv_file(upload_path)
        else:
            sheets = read_excel(upload_path, sheets_mode)

        # 生成TXT内容
        txt_lines = []
        for sheet_name, rows in sheets.items():
            if len(sheets) > 1:
                txt_lines.append(f'========== {sheet_name} ==========')
            txt_lines.extend(rows_to_txt_lines(rows, sep, nl_rep))
            txt_lines.append('')

        txt_content = '\n'.join(txt_lines)

        out_name = uid + '_' + os.path.splitext(filename)[0] + '.txt'
        out_path = os.path.join(app.config['OUTPUT_FOLDER'], out_name)
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(txt_content)

        os.remove(upload_path)

        preview_lines = txt_lines[:120]
        total_lines = len(txt_lines)

        return jsonify({
            'success': True,
            'download_name': out_name,
            'preview': '\n'.join(preview_lines),
            'total_lines': total_lines,
            'sheets': list(sheets.keys())
        })

    except Exception as e:
        return jsonify({'error': f'转换失败: {str(e)}'}), 500

@app.route('/download/<filename>')
def download(filename):
    path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if not os.path.exists(path):
        return '文件不存在', 404
    original_name = '_'.join(filename.split('_')[1:])
    return send_file(path, as_attachment=True, download_name=original_name)

if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('outputs', exist_ok=True)
    print("服务启动中... 请访问 http://127.0.0.1:5000")
    app.run(debug=True, port=5000)
