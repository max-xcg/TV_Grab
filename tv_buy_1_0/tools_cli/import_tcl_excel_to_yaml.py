# -*- coding: utf-8 -*-
"""
tv_buy_1_0/tools_cli/import_tcl_excel_to_yaml.py  （完整版｜可一键复制粘贴替换）

用途：
- 读取 tv_buy_1_0/excel_all 下的 Excel（列=型号、行=参数 的“参数对比表”）
- 输出为 per-model YAML（结构兼容 excel_import_all_v1 目录）
- 支持多品牌（TCL/海信/雷鸟/小米/Vidda/创维），支持 --all_sheets
- 支持去重：同品牌同型号跨 sheet/重复列只写 1 份（并合并字段）
- 关键增强：
  1) 行名模糊匹配（去空白、括号全半角、支持包含/正则）
  2) 递归清理 None：输出 YAML 不再充满 null（未知字段直接省略）
  3) 更宽松解析：价格/亮度/分区等“按尺寸多行”字段可靠提取

示例：
/c/software/Anaconda3/python.exe tv_buy_1_0/tools_cli/import_tcl_excel_to_yaml.py \
  --in_dir tv_buy_1_0/excel_all \
  --out tv_buy_1_0/data_raw/excel_import_all_v1 \
  --overwrite \
  --all_sheets
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Iterable

try:
    import yaml  # type: ignore
except Exception as e:
    raise SystemExit(f"[ERR] missing dependency: pyyaml. Install: pip install pyyaml. detail={e}")

try:
    from openpyxl import load_workbook  # type: ignore
except Exception as e:
    raise SystemExit(f"[ERR] missing dependency: openpyxl. Install: pip install openpyxl. detail={e}")


# =========================
# 品牌识别
# =========================
# 你新增了“创维”，这里加上。
# 说明：
# - “Vidda”仍保持 Vidda（你当前目录就是 Vidda）
# - “创维”目录输出为 创维
_BRAND_HINTS = [
    ("TCL", ["tcl"]),
    ("海信", ["海信", "hisense"]),
    ("雷鸟", ["雷鸟"]),
    ("小米", ["小米", "xiaomi", "mi"]),
    ("Vidda", ["vidda"]),
    ("创维", ["创维", "skyworth"]),
]


def _infer_brand_from_filename(name: str) -> str:
    n = (name or "").strip().lower()
    for brand, kws in _BRAND_HINTS:
        for k in kws:
            if k.lower() in n:
                return brand
    return "Unknown"


def _brand_dir_name(brand: str) -> str:
    """
    输出目录名规范化：
    - 直接返回中文品牌名
    - Vidda 保持 Vidda（你已有目录结构）
    """
    b = (brand or "").strip()
    return b or "Unknown"


def _prefix_for_brand(brand: str) -> str:
    # 文件名统一用品牌前缀
    return _brand_dir_name(brand)


# =========================
# 通用清理：去 None/null
# =========================
def _strip_none(obj: Any) -> Any:
    """
    递归清理 None：
    - dict：删除 value 为 None 的键；若子结构清理后为空，也删除
    - list：删除 None 项；若 dict 清理后为空 dict，则删除该项
    """
    if obj is None:
        return None
    if isinstance(obj, dict):
        out: Dict[str, Any] = {}
        for k, v in obj.items():
            vv = _strip_none(v)
            if vv is None:
                continue
            if isinstance(vv, dict) and not vv:
                continue
            if isinstance(vv, list) and not vv:
                continue
            out[k] = vv
        return out
    if isinstance(obj, list):
        out_list: List[Any] = []
        for v in obj:
            vv = _strip_none(v)
            if vv is None:
                continue
            if isinstance(vv, dict) and not vv:
                continue
            out_list.append(vv)
        return out_list
    return obj


# =========================
# Excel 结构解析
# =========================
def _norm_label(s: str) -> str:
    t = (s or "").strip()
    t = re.sub(r"\s+", "", t)
    t = t.replace("（", "(").replace("）", ")")
    return t


def _collect_label_rows(ws) -> Dict[str, int]:
    label_rows: Dict[str, int] = {}
    max_r = ws.max_row or 0
    for r in range(1, max_r + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        key = _norm_label(s)
        if key and key not in label_rows:
            label_rows[key] = r
    return label_rows


def _find_row_index_exact(label_rows: Dict[str, int], keys: Iterable[str]) -> Optional[int]:
    for k in keys:
        kk = _norm_label(str(k))
        if kk in label_rows:
            return label_rows[kk]
    return None


def _find_row_index_fuzzy(label_rows: Dict[str, int], patterns: Iterable[str]) -> Optional[int]:
    if not label_rows:
        return None
    pats: List[str] = []
    for p in patterns:
        pp = _norm_label(str(p))
        if pp:
            pats.append(pp)

    for label, r in label_rows.items():
        ll = _norm_label(label)
        if not ll:
            continue
        for p in pats:
            if p in ll:
                return r
            try:
                if re.search(p, ll):
                    return r
            except Exception:
                pass
    return None


def _get_cell(ws, label_rows: Dict[str, int], col: int, keys: List[str]) -> Any:
    r = _find_row_index_exact(label_rows, keys)
    if r is None:
        r = _find_row_index_fuzzy(label_rows, keys)
    if r is None:
        return None
    return ws.cell(r, col).value


def _collect_model_cols(ws, label_row: int) -> List[int]:
    cols: List[int] = []
    max_c = ws.max_column or 0
    for c in range(2, max_c + 1):
        v = ws.cell(label_row, c).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        cols.append(c)
    return cols


# =========================
# 字段解析 helpers
# =========================
def _safe_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return int(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(",", "")
    s2 = re.sub(r"[^\d.]", "", s)
    if not s2:
        return None
    try:
        return int(float(s2))
    except Exception:
        return None


def _parse_sizes_list(raw: Any) -> List[int]:
    if raw is None:
        return []
    s = str(raw)
    nums = re.findall(r"(\d{2,3})\s*(?:吋|寸|英寸|\"|inch|in)?", s)
    out: List[int] = []
    for n in nums:
        try:
            v = int(n)
            if 40 <= v <= 120:
                out.append(v)
        except Exception:
            pass
    return sorted(set(out))


def _parse_map_by_size_from_multiline(raw: Any) -> Dict[int, Any]:
    if raw is None:
        return {}
    s = str(raw)
    lines = [x.strip() for x in re.split(r"[\r\n]+", s) if x.strip()]
    out: Dict[int, Any] = {}

    for line in lines:
        m = re.search(r"(\d{2,3})\s*(?:吋|寸|英寸|\"|in|inch)?\s*\|\s*(.+)$", line)
        if not m:
            m = re.search(r"(\d{2,3})\s*(?:吋|寸|英寸|\"|in|inch)?\s*[:： ]\s*(.+)$", line)
        if not m:
            continue

        sz = _safe_int(m.group(1))
        if not sz or not (40 <= sz <= 120):
            continue

        val_raw = (m.group(2) or "").strip()
        v_int = _safe_int(val_raw)
        out[int(sz)] = v_int if v_int is not None else val_raw

    return out


def _parse_refresh_rate(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    s = str(raw).strip()
    if not s:
        return {}
    nums = re.findall(r"(\d{2,3})\s*Hz", s, flags=re.IGNORECASE)
    native = _safe_int(nums[0]) if nums else None
    dynamic = _safe_int(nums[-1]) if len(nums) >= 2 else None
    out: Dict[str, Any] = {"text": s}
    if native is not None:
        out["native_hz"] = native
    if dynamic is not None and dynamic != native:
        out["dynamic_hz"] = dynamic
    return out


def _parse_color_gamut(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    s = str(raw).strip()
    if not s:
        return {}
    m = re.search(r"(\d{2,3})\s*%\s*DCI\-?P3", s, flags=re.IGNORECASE)
    out: Dict[str, Any] = {"text": s}
    if m:
        out["dci_p3_percent"] = _safe_int(m.group(1))
    return out


def _parse_color_accuracy(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    s = str(raw).strip()
    if not s:
        return {}
    out: Dict[str, Any] = {"text": s}
    v = _safe_int(s)
    if v is None:
        try:
            out["delta_e_claim"] = float(re.sub(r"[^\d.]", "", s))
        except Exception:
            pass
    else:
        out["delta_e_claim"] = v
    return out


def _parse_audio(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    s = str(raw).strip()
    if not s:
        return {}
    out: Dict[str, Any] = {"text": s}
    m = re.search(r"(\d+(?:\.\d+){1,2})", s)
    if m:
        out["channels"] = m.group(1)
    elif "双声道" in s:
        out["channels"] = "2.0"
    if "安桥" in s:
        out["brand_or_feature"] = "安桥"
    return out


def _parse_mem_storage(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    s = str(raw).strip()
    if not s:
        return {}
    out: Dict[str, Any] = {"text": s}
    m = re.search(r"(\d+)\s*\+\s*(\d+)\s*GB", s, flags=re.IGNORECASE)
    if m:
        out["ram_gb"] = _safe_int(m.group(1))
        out["storage_gb"] = _safe_int(m.group(2))
    return out


def _parse_hdmi(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    s = str(raw).strip()
    if not s:
        return {}
    out: Dict[str, Any] = {"text": s}

    m21 = re.search(r"HDMI\s*2\.1.*?×\s*(\d+)", s, flags=re.IGNORECASE)
    if m21:
        out["hdmi_2_1_ports"] = _safe_int(m21.group(1))
    m20 = re.search(r"HDMI\s*2\.0.*?×\s*(\d+)", s, flags=re.IGNORECASE)
    if m20:
        out["hdmi_2_0_ports"] = _safe_int(m20.group(1))

    if "满血" in s:
        out["full_bandwidth_2_1"] = True
    elif "非满血" in s:
        out["full_bandwidth_2_1"] = False

    return out


def _parse_usb(raw: Any) -> List[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    lines = [x.strip() for x in re.split(r"[\r\n]+", s) if x.strip()]
    if len(lines) >= 2:
        return lines
    parts = re.split(r"\s{2,}|\s*(?=USB\s*\d)", s)
    out = [p.strip() for p in parts if p.strip()]
    return out or [s]


def _normalize_release(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    return s.replace(".", "-").replace("/", "-")


def _slugify_model(s: str) -> str:
    t = (s or "").strip().lower()
    t = t.replace(" ", "")
    t = t.replace("（", "(").replace("）", ")")
    t = re.sub(r"[^\w\u4e00-\u9fa5\(\)\-]+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t or "model"


def _model_norm_key(model: str) -> str:
    t = (model or "").strip().lower()
    t = t.replace("（", "(").replace("）", ")")
    t = re.sub(r"\s+", "", t)
    t = re.sub(r"[^\w\u4e00-\u9fa5]+", "", t)
    return t


# =========================
# Sheet 解析
# =========================
def _parse_one_sheet(ws, brand: str, sheet_name: str) -> List[Dict[str, Any]]:
    label_rows = _collect_label_rows(ws)

    model_row = _find_row_index_exact(label_rows, ["型号", "机型", "机型/型号"])
    if model_row is None:
        model_row = _find_row_index_fuzzy(label_rows, ["型号", "机型"])
    if model_row is None:
        model_row = 2

    model_cols = _collect_model_cols(ws, label_row=model_row)
    if not model_cols:
        return []

    records: List[Dict[str, Any]] = []

    for col in model_cols:
        model_raw = ws.cell(model_row, col).value
        if model_raw is None:
            continue
        model = str(model_raw).strip()
        if not model:
            continue

        positioning = _get_cell(ws, label_rows, col, ["产品定位", "定位", "产品线"])
        first_release = _get_cell(ws, label_rows, col, ["发布时间", "发布", "上市时间", "首发"])

        display_tech = _get_cell(ws, label_rows, col, ["显示技术", "背光技术", "显示方案"])
        panel = _get_cell(ws, label_rows, col, ["屏幕", "面板", "屏幕类型"])
        picture_chip = _get_cell(ws, label_rows, col, ["画质芯片", "芯片", "画质处理芯片"])

        sizes_raw = _get_cell(ws, label_rows, col, ["包含尺寸", "尺寸", "尺寸覆盖"])
        sizes = _parse_sizes_list(sizes_raw)

        price_sub = _get_cell(ws, label_rows, col, ["国补价预估", "国补价格", "国补价"])
        price_nosub = _get_cell(ws, label_rows, col, ["非国补价预估", "非国补价格", "非国补价"])

        price_85 = _get_cell(
            ws,
            label_rows,
            col,
            [
                "85寸价格（国补预估）",
                "85寸价格(国补预估)",
                "85寸价格（国补）",
                "85寸价格(国补)",
                "85寸价格",
                "85吋价格",
            ],
        )

        brightness_raw = _get_cell(ws, label_rows, col, ["亮度", "峰值亮度", "HDR亮度"])
        zones_raw = _get_cell(ws, label_rows, col, ["分区", "控光分区", "分区数", "背光分区"])

        color_gamut_raw = _get_cell(ws, label_rows, col, ["宣称色域", "色域"])
        color_acc_raw = _get_cell(ws, label_rows, col, ["宣称色准", "色准"])
        refresh_raw = _get_cell(ws, label_rows, col, ["刷新率（Hz）", "刷新率", "刷新率(Hz)", "刷新率hz"])
        audio_raw = _get_cell(ws, label_rows, col, ["音响/扬声器功率", "音响", "扬声器"])
        mem_raw = _get_cell(ws, label_rows, col, ["内存+存储", "内存", "存储"])
        hdmi_raw = _get_cell(ws, label_rows, col, ["HDMI接口", "HDMI", "接口(HDMI)"])
        usb_raw = _get_cell(ws, label_rows, col, ["USB", "usb"])

        price_sub_map = _parse_map_by_size_from_multiline(price_sub)
        price_nosub_map = _parse_map_by_size_from_multiline(price_nosub)

        if price_85 is not None:
            v85 = _safe_int(price_85)
            if v85 is not None:
                price_sub_map.setdefault(85, v85)

        bri_map = _parse_map_by_size_from_multiline(brightness_raw)
        zone_map = _parse_map_by_size_from_multiline(zones_raw)

        all_sizes = list(sizes)
        for m in (price_sub_map, price_nosub_map, bri_map, zone_map):
            for k in m.keys():
                if k not in all_sizes:
                    all_sizes.append(k)
        all_sizes = sorted(set([int(x) for x in all_sizes if 40 <= int(x) <= 120]))

        variants: List[Dict[str, Any]] = []
        for sz in all_sizes:
            v: Dict[str, Any] = {"size_inch": int(sz)}
            if sz in price_sub_map:
                v["price_cny"] = price_sub_map.get(sz)
            if sz in price_nosub_map:
                v["price_before_subsidy_cny"] = price_nosub_map.get(sz)
            if sz in bri_map:
                v["peak_brightness_nits"] = bri_map.get(sz)
            if sz in zone_map:
                v["dimming_zones"] = zone_map.get(sz)
            variants.append(_strip_none(v))

        spec: Dict[str, Any] = {}
        if display_tech is not None:
            spec["display_tech"] = str(display_tech).strip()
        if panel is not None:
            spec["panel"] = str(panel).strip()
        if picture_chip is not None:
            spec["picture_chip"] = str(picture_chip).strip()

        cg = _parse_color_gamut(color_gamut_raw)
        if cg:
            spec["color_gamut"] = cg
        ca = _parse_color_accuracy(color_acc_raw)
        if ca:
            spec["color_accuracy"] = ca
        rr = _parse_refresh_rate(refresh_raw)
        if rr:
            spec["refresh_rate"] = rr
        au = _parse_audio(audio_raw)
        if au:
            spec["audio"] = au
        ms = _parse_mem_storage(mem_raw)
        if ms:
            spec["memory_storage"] = ms
        hd = _parse_hdmi(hdmi_raw)
        if hd:
            spec["hdmi"] = hd
        usb = _parse_usb(usb_raw)
        if usb:
            spec["usb"] = usb

        obj: Dict[str, Any] = {
            "brand": brand,
            "model": model,
            "first_release": _normalize_release(first_release),
            "positioning": str(positioning).strip() if positioning is not None else None,
            "display_tech": str(display_tech).strip() if display_tech is not None else None,
            "panel": str(panel).strip() if panel is not None else None,
            "picture_chip": str(picture_chip).strip() if picture_chip is not None else None,
            "spec": spec or None,
            "variants": variants,
            "source": {
                "kind": "excel",
                "file": None,
                "sheet": sheet_name,
                "column_model": model,
            },
        }

        records.append(_strip_none(obj))

    return records


# =========================
# 去重合并（同品牌同型号）
# =========================
def _merge_variants(old: List[Dict[str, Any]], new: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    mp: Dict[int, Dict[str, Any]] = {}
    for it in old or []:
        sz = _safe_int(it.get("size_inch"))
        if not sz:
            continue
        mp[int(sz)] = dict(it)

    for it in new or []:
        sz = _safe_int(it.get("size_inch"))
        if not sz:
            continue
        cur = mp.get(int(sz), {})
        merged = dict(cur)
        for k, v in it.items():
            if v is None:
                continue
            if k not in merged or merged.get(k) in (None, "", []):
                merged[k] = v
        mp[int(sz)] = merged

    out = [_strip_none(x) for x in mp.values()]
    out = sorted(out, key=lambda x: int(x.get("size_inch") or 0))
    return out


def _merge_records(old: Dict[str, Any], new: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(old)
    for k, v in new.items():
        if k == "variants":
            continue
        if v is None:
            continue
        if k not in out or out.get(k) in (None, "", []):
            out[k] = v

    # spec 合并
    if isinstance(old.get("spec"), dict) or isinstance(new.get("spec"), dict):
        spec = dict(old.get("spec") or {})
        for sk, sv in (new.get("spec") or {}).items():
            if sv is None:
                continue
            if sk not in spec or spec.get(sk) in (None, "", []):
                spec[sk] = sv
        out["spec"] = spec or None

    out["variants"] = _merge_variants(old.get("variants") or [], new.get("variants") or [])

    # source 合并
    if isinstance(old.get("source"), dict) and isinstance(new.get("source"), dict):
        src = dict(old["source"])
        for sk, sv in new["source"].items():
            if sv is None:
                continue
            if sk not in src or src.get(sk) in (None, "", []):
                src[sk] = sv
        out["source"] = src

    return _strip_none(out)


def _dedup_records(records: List[Dict[str, Any]], brand: str) -> Tuple[List[Dict[str, Any]], int]:
    mp: Dict[str, Dict[str, Any]] = {}
    skipped = 0
    for obj in records:
        model = str(obj.get("model") or "").strip()
        if not model:
            continue
        key = f"{_brand_dir_name(brand).lower()}::{_model_norm_key(model)}"
        if key in mp:
            mp[key] = _merge_records(mp[key], obj)
            skipped += 1
        else:
            mp[key] = obj
    return list(mp.values()), skipped


# =========================
# YAML 写入
# =========================
def _dump_yaml(obj: Dict[str, Any]) -> str:
    cleaned = _strip_none(obj)
    return yaml.safe_dump(cleaned, allow_unicode=True, sort_keys=False)


def _write_yaml(path: Path, obj: Dict[str, Any], overwrite: bool) -> bool:
    if path.exists() and not overwrite:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(_dump_yaml(obj), encoding="utf-8")
    return True


# =========================
# 主流程
# =========================
def convert_excel_file(
    xlsx_path: Path,
    out_root: Path,
    dry_run: bool,
    limit: Optional[int],
    overwrite: bool,
    sheet: Optional[str],
    all_sheets: bool,
    brand_override: Optional[str],
) -> Tuple[str, int, int, int, int, int]:
    """
    return: (brand, parsed_total, unique_models, written_unique, skipped_duplicates, sheet_cnt)
    """
    brand = (brand_override or _infer_brand_from_filename(xlsx_path.name)).strip() or "Unknown"

    wb = load_workbook(xlsx_path, data_only=True)

    sheet_names = list(wb.sheetnames)
    if sheet:
        if sheet not in sheet_names:
            raise SystemExit(f"[ERR] sheet not found in {xlsx_path.name}: {sheet}. available={sheet_names}")
        target_sheets = [sheet]
    elif all_sheets:
        target_sheets = sheet_names
    else:
        target_sheets = [sheet_names[0]]

    parsed_total = 0
    all_records: List[Dict[str, Any]] = []

    for sn in target_sheets:
        ws = wb[sn]
        recs = _parse_one_sheet(ws, brand=brand, sheet_name=sn) or []
        for r in recs:
            if isinstance(r.get("source"), dict):
                r["source"]["file"] = xlsx_path.name
        if limit is not None:
            recs = recs[: int(limit)]
        parsed_total += len(recs)
        all_records.extend(recs)

    unique_records, skipped_dup = _dedup_records(all_records, brand=brand)
    unique_models = len(unique_records)

    out_dir = out_root / _brand_dir_name(brand)
    written_unique = 0

    for obj in unique_records:
        model = obj.get("model") or "model"
        prefix = _prefix_for_brand(brand)
        slug = _slugify_model(str(model))
        out_name = f"{prefix}_{slug}.yaml"
        out_path = out_dir / out_name

        if dry_run:
            written_unique += 1
            continue

        if _write_yaml(out_path, obj, overwrite=overwrite):
            written_unique += 1

    return brand, parsed_total, unique_models, written_unique, skipped_dup, len(target_sheets)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in_dir", required=True, help="Input excel dir, e.g. tv_buy_1_0/excel_all")
    ap.add_argument("--out", default="tv_buy_1_0/data_raw/excel_import_all_v1", help="Output yaml root dir")
    ap.add_argument("--dry_run", action="store_true", help="Parse only, do not write files")
    ap.add_argument("--limit", type=int, default=None, help="Limit models per excel (debug)")
    ap.add_argument("--overwrite", action="store_true", help="Overwrite existing yaml files")
    ap.add_argument("--sheet", default=None, help="Only parse a specific sheet name")
    ap.add_argument("--all_sheets", action="store_true", help="Parse all sheets (recommended)")
    ap.add_argument("--brand", default=None, help="Override brand for this run (optional)")
    args = ap.parse_args()

    in_dir = Path(args.in_dir)
    out_root = Path(args.out)

    if not in_dir.exists() or not in_dir.is_dir():
        raise SystemExit(f"[ERR] in_dir not found: {in_dir}")

    excel_files = sorted([p for p in in_dir.iterdir() if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm")])
    if not excel_files:
        raise SystemExit(f"[ERR] no excel files in: {in_dir}")

    total_excel = 0
    total_parsed = 0
    total_unique = 0
    total_written = 0
    total_skipped = 0

    for xlsx in excel_files:
        total_excel += 1
        brand, parsed_total, unique_models, written_unique, skipped_dup, sheet_cnt = convert_excel_file(
            xlsx_path=xlsx,
            out_root=out_root,
            dry_run=bool(args.dry_run),
            limit=args.limit,
            overwrite=bool(args.overwrite),
            sheet=args.sheet,
            all_sheets=bool(args.all_sheets),
            brand_override=args.brand,
        )

        total_parsed += parsed_total
        total_unique += unique_models
        total_written += written_unique
        total_skipped += skipped_dup

        print(
            f"[IN ] {xlsx.name}  brand={brand}  sheets={sheet_cnt}  parsed_total={parsed_total}  "
            f"unique_models={unique_models}  written_unique={written_unique}  "
            f"skipped_duplicates={skipped_dup}  out={out_root / _brand_dir_name(brand)}"
        )

    print(
        f"[DONE] excel_files={total_excel} parsed_total={total_parsed} unique_models={total_unique} "
        f"written_unique={total_written} skipped_duplicates={total_skipped}"
    )


if __name__ == "__main__":
    main()